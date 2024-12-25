#!/usr/bin/env python3

import sys
try:
   from openpyxl import load_workbook
   from fpdf import FPDF
except Exception as e:
   print(e)
   sys.exit(1)

import argparse
from pathlib import Path

from typing import NamedTuple  #not to be confused with namedtuple in collections
class Caller_lists(NamedTuple):
    success: bool = False
    message: str = ''
    caller_mapping_dict: dict = {}
    no_guest_list: list = []
    guest_dict: dict = {}

def make_guests_per_caller_lists(in_filename):
   # returns the tuple Caller_lists

   Caller_lists()
   Caller_lists.success = False # default value didn't seem to work

   try:
      workbook = load_workbook(in_filename, data_only=True)
   except Exception as e:
      Caller_lists.message = f"Error when reading {in_filename}: {e}"
      return Caller_lists
   
   sheetnames = workbook.sheetnames
   # print(f"{workbook.sheetnames=}")

   EXPECTED_SHEETNAMES = ['guest-to-caller', 'callers', 'guests']
   if sheetnames != EXPECTED_SHEETNAMES:
      Caller_lists.message = f"Error: expected '{EXPECTED_SHEETNAMES}' sheet names; found '{workbook.sheetnames}' in file '{in_filename}'"
      return Caller_lists

   #make dictionary of caller, [guests]
   mapping_dict = {}
   is_header = True
   for row in workbook['callers'].rows:
      # skip header row; there is only 1 column: the list of callers
      if is_header:
         is_header = False
      else:
         mapping_dict[row[0].value] = []
   
   is_header = True
   for row in workbook['guest-to-caller'].rows:
      # columns: 0=Guest, 1=Caller, 2=Note
      if is_header:
         is_header = False
      else:
         # a list containing the guest name and note to the caller's list:
         mapping_dict[row[1].value].append([row[0].value, row[2].value])
         # print(f'{row[0].value} -> {row[1].value}')
   # print(f"{mapping_dict=}\n")
   '''
   mapping_dict={'Caroline': [['Guest1', 'new regular']], 'Tina': [], 'Peter': [], 'Rebecca': [['Guest2', 'Substitute this week only']], 'Maria': [], 'Barb': [], 'Lisa': [['Guest3', None]], 'Do-Not-Call': []}
   '''
   callers_with_no_guests = []
   for caller, guests in mapping_dict.items():
      if len(guests) == 0:
         callers_with_no_guests.append(caller)
   # remove caller with no guests from mapping_dict
   for caller in callers_with_no_guests:   
         mapping_dict.pop(caller) 

   # make a dictionary of guest data to be used for generating reports.
   guest_dict = {}
   is_header = True
   for row in workbook['guests'].rows:
      # columns: 0=First, 1=Last, 2=UserName, 3=Password, 4=Town, 5=Phone, 6=Notes
      if is_header:
         is_header = False
      else:
         guest_dict[row[2].value]= {'First':row[0].value, 'Last':row[1].value, 'PW':row[3].value,
                                 'Town':row[4].value, 'Phone':row[5].value, 'Notes':row[6].value}      
   # print(f"{guest_dict=}")
   '''
   guest_dict={'Guest1': {'First': 'Guest', 'Last': 1.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call early'}, 'Guest2': {'First': 'Guest', 'Last': 2.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call 3 times'}, 'Guest3': {'First': 'Guest', 'Last': 3.0, 'PW': 'secret', 'Town': 'Newbury', 'Phone': '978.555.0000', 'Notes': 'call late'}}
   '''
   
   Caller_lists.no_guest_list = callers_with_no_guests
   Caller_lists.caller_mapping_dict = mapping_dict
   Caller_lists.guest_dict = guest_dict
   Caller_lists.success = True

   return Caller_lists


def make_caller_pdfs(caller_mapping_dict, guest_dict):
   # PDF writing example: https://medium.com/@mahijain9211/creating-a-python-class-for-generating-pdf-tables-from-a-pandas-dataframe-using-fpdf2-c0eb4b88355c
   for caller, guests in caller_mapping_dict.items():
      pdf = FPDF(orientation="L", format="letter") # default units are mm; adding , unit="in" inserts blank pages
      pdf.add_page()
      pdf.set_font("Helvetica", size=14)
      pdf.cell(0, 10, f"{caller} - Friday, Dec 27", align="C")
      pdf.ln(10)

      with pdf.table(col_widths=(11,13,14,13,13,14,14,30), line_height=6) as table:
         pdf.set_font(size=12)
         header = ['First', 'Last', 'UserName', 'Password', 'Town', 'Phone', 'Caller notes', 'Notes about guest']
         row = table.row()
         for column in header:
            row.cell(column)

         for guest_id_note in guests:
            this_guest = guest_dict[guest_id_note[0]]
            this_weeks_guest_note = guest_id_note[1]
            if this_weeks_guest_note is None:
               this_weeks_guest_note = ''
            row_data = [this_guest['First'], this_guest['Last'], guest_id_note[0], \
                        this_guest['PW'], this_guest['Town'], this_guest['Phone'], \
                        this_weeks_guest_note, this_guest['Notes']]
            row = table.row()
            for item in row_data:
               row.cell(str(item), v_align="T")

      pdf.output(f"{caller}.pdf")
      # break

if __name__ == "__main__":
   argParser = argparse.ArgumentParser()
   argParser.add_argument("input", type=str, help="input filename with path")

   args = argParser.parse_args()
   # print(f'\n\t{args.input= } {args.auth_path= }\n\t{args.dont_email= } {args.parse_only= }')

   if args.input is None:
      sys.exit("No file selected to parse.")
   elif Path(args.input).is_file():
      filename = args.input
   else:
      sys.exit("file selected is not a file.")

   Caller_lists = make_guests_per_caller_lists(filename)
   if not Caller_lists.success:
      print(f"Failure: {Caller_lists.message}")
      sys.exit(1)

   make_caller_pdfs(Caller_lists.caller_mapping_dict, Caller_lists.guest_dict)
   
   # print(f"Callers with no guests: {Caller_lists.no_guest_list}")
   # print(f"Guest info: {Caller_lists.guest_dict}")
   # print(f"Guest per caller: {Caller_lists.caller_mapping_dict}")
